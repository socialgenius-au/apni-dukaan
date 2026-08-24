"""
Export a fillable PRICE LIST template for Hamro Pashupatinath Mart (merchant_id=7).
Produces an .xlsx with: Product Name, Category, Unit, Current Price, and a
blank highlighted "New Price" column for the merchant to fill in.
"""

import os
import requests
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

API_URL = "https://hamari-dukaan-production.up.railway.app"
MERCHANT_ID = 7
MERCHANT_NAME = "Hamro Pashupatinath Mart"

def main():
    url = f"{API_URL}/products/merchant/{MERCHANT_ID}/all"
    print(f"Fetching from {url} ...")
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    products = resp.json()
    print(f"Fetched {len(products)} products for {MERCHANT_NAME}")

    rows = []
    for p in products:
        rows.append({
            "Product Name": p.get("name"),
            "Category": p.get("category"),
            "Unit": p.get("unit"),
            "Current Price": p.get("price"),
            "New Price": None,
        })

    df = pd.DataFrame(rows).sort_values(by=["Category", "Product Name"], na_position="last")

    out_path = "pashupatinath_price_list.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Price List", startrow=1)
        ws = writer.sheets["Price List"]

        ws["A1"] = "Fill in the yellow 'New Price' column only. Leave blank to keep current price."
        ws["A1"].font = Font(italic=True, size=10, color="555555")

        header_row = 2
        n_cols = len(df.columns)
        n_rows = len(df)

        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = Font(bold=True, name="Arial")
            cell.alignment = Alignment(horizontal="center")

        new_price_col = n_cols
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for r in range(header_row + 1, header_row + 1 + n_rows):
            ws.cell(row=r, column=new_price_col).fill = yellow

        for row in ws.iter_rows(min_row=header_row, max_row=header_row + n_rows, max_col=n_cols):
            for cell in row:
                if cell.font.name != "Arial":
                    cell.font = Font(name="Arial", bold=cell.font.bold)

        widths = {1: 40, 2: 18, 3: 10, 4: 14, 5: 14}
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    print(f"Saved: {out_path}")
    print(f"Total products: {len(df)}")

if __name__ == "__main__":
    main()
